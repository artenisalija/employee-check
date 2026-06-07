from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import STATUS_CHECKED_IN, STATUS_CHECKED_OUT, STATUS_LUNCH, STATUS_MEETING
from .paths import reports_dir
from .storage import EmployerStore


def generate_daily_report(store: EmployerStore, report_date: datetime | None = None, output_dir: Path | None = None) -> Path:
    day = (report_date or datetime.now(timezone.utc)).astimezone().date()
    local_start = datetime.combine(day, datetime.min.time()).astimezone()
    local_end = local_start + timedelta(days=1)
    start_iso = local_start.astimezone(timezone.utc).replace(microsecond=0).isoformat()
    end_iso = local_end.astimezone(timezone.utc).replace(microsecond=0).isoformat()
    rows = store.snapshots_between(start_iso, end_iso)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    apps_ws = wb.create_sheet("Per App")
    raw_ws = wb.create_sheet("Raw")

    summary = _summarize(rows)
    _write_summary(summary_ws, summary, day.isoformat())
    _write_apps(apps_ws, summary)
    _write_raw(raw_ws, rows)

    target_dir = output_dir or reports_dir()
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / f"employee-summary-{day.isoformat()}.xlsx"
    wb.save(path)
    return path


def _summarize(rows: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    by_employee: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_employee[(row["employee_name"], row["machine_name"])].append(row)

    for key, employee_rows in by_employee.items():
        employee_rows.sort(key=lambda item: item["timestamp"])
        active_seconds = 0.0
        idle_seconds = 0.0
        checked_in_seconds = 0.0
        lunch_seconds = 0.0
        meeting_seconds = 0.0
        checked_out_seconds = 0.0
        app_seconds: dict[str, float] = defaultdict(float)
        first_seen = employee_rows[0]["timestamp"]
        last_seen = employee_rows[-1]["timestamp"]
        current = employee_rows[-1]

        for index, row in enumerate(employee_rows):
            if index + 1 < len(employee_rows):
                delta = _seconds_between(row["timestamp"], employee_rows[index + 1]["timestamp"])
            else:
                delta = 0.0
            delta = min(max(delta, 0.0), 60.0)
            status = row["manual_status"]
            if status == STATUS_CHECKED_IN:
                checked_in_seconds += delta
            if status == STATUS_CHECKED_OUT:
                checked_out_seconds += delta
            elif status == STATUS_LUNCH:
                lunch_seconds += delta
            elif status == STATUS_MEETING:
                meeting_seconds += delta
            elif float(row["idle_seconds"] or 0) >= 120:
                idle_seconds += delta
            else:
                active_seconds += delta
                app = row.get("active_app") or row.get("active_process") or "Unknown"
                app_seconds[app] += delta

        grouped[key] = {
            "employee_name": key[0],
            "machine_name": key[1],
            "first_seen": employee_rows[0].get("local_timestamp") or first_seen,
            "last_seen": employee_rows[-1].get("local_timestamp") or last_seen,
            "active_seconds": active_seconds,
            "idle_seconds": idle_seconds,
            "checked_in_seconds": checked_in_seconds,
            "lunch_seconds": lunch_seconds,
            "meeting_seconds": meeting_seconds,
            "checked_out_seconds": checked_out_seconds,
            "current_status": current["manual_status"],
            "current_status_elapsed_seconds": float(current.get("status_elapsed_seconds") or 0),
            "current_idle_band": current["idle_band"],
            "current_app": current.get("active_app") or current.get("active_process") or "",
            "current_title": current.get("active_title") or "",
            "current_url": current.get("active_url") or "",
            "current_machine_time": current.get("local_timestamp") or current.get("timestamp") or "",
            "app_seconds": dict(app_seconds),
        }
    return grouped


def _write_summary(ws, summary: dict[tuple[str, str], dict[str, Any]], report_day: str) -> None:
    ws.append([f"Employee Check Summary - {report_day}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws["A1"].font = Font(bold=True, size=14)
    headers = [
        "Employee",
        "Machine",
        "First Seen",
        "Last Seen",
        "Active",
        "Idle",
        "Checked In Status",
        "Lunch",
        "Meeting",
        "Checked Out",
        "Current Status",
        "Current Status Time",
        "Current Idle Band",
        "Current App",
        "Current Machine Time",
        "Current Title / URL",
    ]
    ws.append(headers)
    _style_header(ws, 2)
    for item in summary.values():
        title_or_url = item["current_url"] or item["current_title"]
        ws.append(
            [
                item["employee_name"],
                item["machine_name"],
                item["first_seen"],
                item["last_seen"],
                _format_seconds(item["active_seconds"]),
                _format_seconds(item["idle_seconds"]),
                _format_seconds(item["checked_in_seconds"]),
                _format_seconds(item["lunch_seconds"]),
                _format_seconds(item["meeting_seconds"]),
                _format_seconds(item["checked_out_seconds"]),
                item["current_status"],
                _format_seconds(item["current_status_elapsed_seconds"]),
                item["current_idle_band"],
                item["current_app"],
                item["current_machine_time"],
                title_or_url,
            ]
        )
    _autosize(ws)


def _write_apps(ws, summary: dict[tuple[str, str], dict[str, Any]]) -> None:
    ws.append(["Employee", "Machine", "App", "Active Time"])
    _style_header(ws, 1)
    for item in summary.values():
        for app, seconds in sorted(item["app_seconds"].items(), key=lambda pair: pair[1], reverse=True):
            ws.append([item["employee_name"], item["machine_name"], app, _format_seconds(seconds)])
    _autosize(ws)


def _write_raw(ws, rows: list[dict[str, Any]]) -> None:
    headers = [
        "Timestamp",
        "Machine Local Timestamp",
        "Employee",
        "Machine",
        "Manual Status",
        "Status Started At",
        "Status Elapsed Seconds",
        "Status Totals JSON",
        "Status Totals Day",
        "Idle Seconds",
        "Idle Band",
        "Active App",
        "Active Process",
        "Active Title",
        "Active URL",
        "Open Apps JSON",
    ]
    ws.append(headers)
    _style_header(ws, 1)
    for row in rows:
        ws.append(
            [
                row["timestamp"],
                row.get("local_timestamp") or "",
                row["employee_name"],
                row["machine_name"],
                row["manual_status"],
                row.get("status_started_at") or "",
                row.get("status_elapsed_seconds") or 0,
                row.get("status_totals_json") or "{}",
                row.get("status_totals_day") or "",
                row["idle_seconds"],
                row["idle_band"],
                row.get("active_app") or "",
                row.get("active_process") or "",
                row.get("active_title") or "",
                row.get("active_url") or "",
                row.get("open_apps_json") or "[]",
            ]
        )
    _autosize(ws)


def _seconds_between(start_iso: str, end_iso: str) -> float:
    try:
        start = datetime.fromisoformat(start_iso)
        end = datetime.fromisoformat(end_iso)
    except ValueError:
        return 0.0
    return max(0.0, (end - start).total_seconds())


def _format_seconds(seconds: float) -> str:
    seconds = int(max(0, seconds))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def _style_header(ws, row: int) -> None:
    fill = PatternFill("solid", fgColor="E8EEF7")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top")


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(value), 80))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[letter].width = max(12, min(length + 2, 80))
